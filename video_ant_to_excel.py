#!/usr/local/bin/python3

import sys
import json

from openpyxl import Workbook, load_workbook

def getJsonData(json_file_path):
    with open(json_file_path, 'r') as f:
        return json.load(f)


def editExcel(workbook, json_data):
    wb = load_workbook(workbook)
    ws = wb.active

    ws['A1'] = json_data['ant']['title']

    ws['A3'] = 'Time stamp of post'
    ws['B3'] = 'Name of poster'
    ws['C3'] = 'Subject of post'
    ws['D3'] = 'Text of initial post'

    current_row = 4
    for annotation in json_data['ant']['annotations']:
        ws['A{}'.format(current_row)] = annotation['annotation']['seconds']
        ws['B{}'.format(current_row)] = annotation['annotation']['author']
        ws['C{}'.format(current_row)] = annotation['annotation']['subject']
        ws['D{}'.format(current_row)] = annotation['annotation']['content']

        current_column = 'E'
        comment_number = 1

        for comment in annotation['annotation']['comments']:
            ws['{}3'.format(current_column)] = 'Comment {}'.format(comment_number)
            ws['{}{}'.format(current_column, current_row)] = comment['comment']['content']

            current_column = chr(ord(current_column) + 1) #https://www.geeksforgeeks.org/ways-increment-character-python/

            ws['{}3'.format(current_column)] = 'Comment {} Author'.format(comment_number)
            ws['{}{}'.format(current_column, current_row)] = comment['comment']['author']

            comment_number += 1

        current_row += 1


    wb.save(workbook)


if __name__ == '__main__':
    videoant_json_file = sys.argv[1]

    json_data = getJsonData(videoant_json_file)

    workbook = sys.argv[2]

    editExcel(workbook, json_data)


