#!/usr/local/bin/python3
import requests
from openpyxl import Workbook

def getAlbumVideos(album_id):
    url = 'https://api.vimeo.com/me/albums/{}/videos?fields=uri,name'.format(album_id)
    headers = {'Authorization': 'Bearer 541616e8df75120c3a35669892dc4b59', 'Content-Type':'application/vnd.vimeo.*+json;version=3.4'}
    r = requests.get(url,headers=headers)
    json = r.json()

    album_videos = []
    for video in json['data']:
        print(video)
        album_videos.append((video['uri'], video['name']))

    return album_videos


if __name__ == '__main__':
    albums = {
        'Bobby Gagnon' :'5486038',
        'Reinsvold':'5483866',
        'Docs':'5439433',
        'FieldScope Tutorials':'5397824',
        'VATL Canvas':'5336013',
        'STeLLA CO2':'5227424',
        '3DMSS PD':'5182319',
        'EMAT':'4676400',
        'ViSTA Plus':'4583884',
        '3DMSS Canvas':'4388449',
        'MNSTL':'4379454',
        'Allergies':'4259252'
    }

    current_line = 1
    wb = Workbook()
    ws = wb.active

    for album in albums:
        album_videos = getAlbumVideos(albums[album])

        ws['A{}'.format(current_line)] = album
        current_line += 1
        for video in album_videos:
            ws['B{}'.format(current_line)] = video[0]
            ws['C{}'.format(current_line)] = video[1]
            current_line += 1


    wb.save('vimeo_transfer_list.xlsx')
