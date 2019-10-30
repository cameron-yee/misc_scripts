#!/usr/local/bin/python3
from openpyxl import Workbook, load_workbook
from datetime import date, timedelta
import sys
import re
#from tkinter import *

def inputPercentage():
    percentage = input('Enter percentage: ')
    m = re.search('[a-zA-Z]', percentage)
    if m is not None:
        print('Enter valid percentage')
        return inputPercentage()
    return percentage


def inputProjectCode():
    inp = input('Enter project code: ')
    code = inp.lower()
    if code not in codes:
        print('Not a valid code')
        return inputProjectCode()

    return code


def getTimePeriod():
    today = date.today().strftime('%m-%d-%y')
    return today


def getPayPeriod():
    pay_periods = {
        '03-24-19': '=S15',
        '04-07-19': '=S16',
        '04-21-19': '=S17',
        '05-05-19': '=S18',
        '05-19-19': '=S19',
        '06-02-19': '=S20',
        '06-16-19': '=S21',
        '06-30-19': '=S22',
        '07-14-19': '=S23',
        '07-28-19': '=S24',
        '08-11-19': '=S25',
        '08-25-19': '=S26',
        '09-08-19': '=S27',
        '09-22-19': '=S28',
        '10-06-19': '=S29',
        '10-20-19': '=S30',
        '11-03-19': '=S31',
        '11-17-19': '=S32',
        '12-01-19': '=S33'
    }

    today = getTimePeriod()

    previous_pay_period = '06-17-18'
    cell_value = None
    pay_period = None

    for key, value in sorted(pay_periods.items()):
        if today < key and today > previous_pay_period:
            pay_period = previous_pay_period #date as str
            cell_value = pay_periods[previous_pay_period] #value of last dict key
            return pay_period, cell_value

        previous_pay_period = key


def getTurnInDate(pay_period_date):
    m = re.findall('([0-9]+)', pay_period_date)
    print(m)
    month = m[0]
    day = m[1]
    year = '20{}'.format(m[2])
    turn_in_date = (date(int(year), int(month), int(day)) + timedelta(days=13)).strftime('%m-%d-%y')

    return turn_in_date


def editExcelStaticValues():
    pay_period = getPayPeriod()

    turn_in_date = getTurnInDate(pay_period[0])

    try:
        hours_excel = load_workbook('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))
    except:
        hours_excel = clearWorkbook()

    ws = hours_excel.active

    ws['E7'] = pay_period[1]
    ws['B2'] = 'Cameron Yee'

    hours_excel.save('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))


def editExcel(workbook, percentage, code, turn_in_date):
    hours_excel = workbook
    ws = hours_excel.active

    #Gets active projects and sorts alphabetically
    def sortProjects():
        count = 0
        active_projects = {}

        for r in range(12,28):
            if ws['C{}'.format(r)].value is not None:
                active_projects[ws['B{}'.format(r)].value] = ws['C{}'.format(r)].value
                count += 1

        if code.upper() in active_projects:
            active_projects[code.upper()] += percentage #(float(hours)/float(80)*float(100))
        else:
            active_projects[code.upper()] = percentage #(float(hours)/float(80)*float(100))

        sorted_projects = sorted(active_projects.items(), key=lambda a: a[0])
        return sorted_projects

    sorted_projects = sortProjects()

    #TODO: Refactor 3 for loops into 1-2
    #Alphabetically edits rows with active projects and hours
    row=12
    for pair in sorted_projects: 
        ws['B{}'.format(row)].value = pair[0]
        ws['C{}'.format(row)].value = float(pair[1])
        row += 1

    #Sets values for all empty rows
    for i in range(row, 28):
        ws['B{}'.format(i)].value = 'Active Project'
        ws['C{}'.format(i)].value = None

    hours_excel.save('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))


def addPercentage(percentage, code):
    pay_period = getPayPeriod()

    turn_in_date = getTurnInDate(pay_period[0])

    try:
        hours_excel = load_workbook('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))
    except:
        hours_excel = clearWorkbook()

    editExcel(hours_excel, percentage, code, turn_in_date)


def addPTO(percentage):
    pay_period = getPayPeriod()

    turn_in_date = getTurnInDate(pay_period[0])

    try:
        hours_excel = load_workbook('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))
    except:
        hours_excel = clearWorkbook()

    ws = hours_excel.active
    ws['C11'] = float(percentage)

    hours_excel.save('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))


def clearWorkbook():
    pay_period = getPayPeriod()[0]
    turn_in_date = getTurnInDate(pay_period)
    hours_excel = load_workbook('/Users/cyee/Documents/Timesheets/template.xlsx')
    hours_excel.save('/Users/cyee/Documents/Timesheets/Timesheet_Cyee_{}.xlsx'.format(turn_in_date))
    return hours_excel


def loopScript(clear=None, first=True):
    try:
        if clear == 'clear':
            clearWorkbook()

        # Only call this once
        if first:
            editExcelStaticValues()

        percentage = inputPercentage()
        code = inputProjectCode()

        addPercentage(percentage, code) if code != 'pto' else addPTO(percentage)

        loopScript(None, False)
    except KeyboardInterrupt:
        print('\nGoodbye Sir')


if __name__ == '__main__':
    codes = ['comm','web','prod','3dmss','hlit','psspt','vatl','prch1', 'stlon', 'pto']

    try:
        loopScript(sys.argv[1]) if len(sys.argv) > 1 else loopScript()
    except KeyboardInterrupt:
        print('\nGoodbye Sir')
