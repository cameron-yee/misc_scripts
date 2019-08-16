#!/usr/local/bin/python3

from openpyxl import Workbook, load_workbook

def getEmailList(workbook):
    wb = load_workbook(workbook)
    ws = wb.active

    email_excel = '/Users/cyee/Desktop/emails.xlsx'

    new_wb = load_workbook(email_excel)
    new_ws = new_wb.active

    existing_emails = []

    row = 1
    for i in range(1, 3142):
        email = ws['E{}'.format(i)].value

        if email not in existing_emails and email is not None:
            new_ws['A{}'.format(row)] = email
            row +=1

            existing_emails.append(email)

    new_wb.save(email_excel)


if __name__ == '__main__':
    getEmailList('/Users/cyee/Dropbox/MSSci Registrations/mssci USER LIST UPDATED.xlsx')


